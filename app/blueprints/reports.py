"""Reports Blueprint — Excel & PDF export."""

from datetime import date, timedelta
from flask import Blueprint, render_template, request, send_file, redirect, url_for
from ..models import db, Member, MembershipPlan, Payment, Settings
from ..blueprints.auth import login_required
import io

reports_bp = Blueprint('reports', __name__, url_prefix='/reports')


@reports_bp.route('/')
@login_required
def index():
    settings = Settings.query.first()
    today = date.today()

    # Revenue per month (last 12)
    monthly = []
    for i in range(11, -1, -1):
        d = (today.replace(day=1) - timedelta(days=i*30))
        label = d.strftime('%b %Y')
        rev = db.session.query(db.func.sum(Member.amount_paid)).filter(
            db.func.strftime('%Y-%m', Member.joining_date) == d.strftime('%Y-%m')
        ).scalar() or 0
        monthly.append({'label': label, 'revenue': float(rev)})

    total_revenue = db.session.query(db.func.sum(Member.amount_paid)).scalar() or 0
    total_due = db.session.query(db.func.sum(Member.total_fee - Member.amount_paid)).scalar() or 0

    return render_template('reports/index.html', settings=settings, monthly=monthly,
                           total_revenue=total_revenue, total_due=total_due)


@reports_bp.route('/export/excel')
@login_required
def export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    settings = Settings.query.first()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Members'

    # Header
    headers = ['Member ID', 'Name', 'Mobile', 'Gender', 'Age', 'Plan', 'Category',
               'Joining Date', 'Expiry Date', 'Total Fee', 'Paid', 'Due', 'Status']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='0A1628')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 16

    members = Member.query.order_by(Member.joining_date.desc()).all()
    for row, m in enumerate(members, 2):
        ws.append([
            m.member_id, m.full_name, m.mobile, m.gender, m.age,
            m.plan.name if m.plan else '', m.plan.category if m.plan else '',
            m.joining_date.isoformat(), m.expiry_date.isoformat(),
            m.total_fee, m.amount_paid, m.due_amount, m.status
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f'{settings.gym_name}_members.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@reports_bp.route('/export/pdf')
@login_required
def export_pdf():
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    settings = Settings.query.first()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=20, rightMargin=20,
                             topMargin=30, bottomMargin=20)

    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph(f"{settings.gym_name} — Member Report", styles['Title']))
    story.append(Paragraph(f"Generated: {date.today().isoformat()}", styles['Normal']))
    story.append(Spacer(1, 12))

    headers = ['ID', 'Name', 'Mobile', 'Plan', 'Joining', 'Expiry', 'Paid', 'Due', 'Status']
    data = [headers]
    members = Member.query.order_by(Member.joining_date.desc()).all()
    for m in members:
        data.append([
            m.member_id, m.full_name[:20], m.mobile,
            f"{m.plan.name} ({m.plan.category})" if m.plan else '',
            m.joining_date.isoformat(), m.expiry_date.isoformat(),
            f"₹{m.amount_paid:.0f}", f"₹{m.due_amount:.0f}", m.status
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0A1628')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))
    story.append(table)
    doc.build(story)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f'{settings.gym_name}_members.pdf',
                     mimetype='application/pdf')
